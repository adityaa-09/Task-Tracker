import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime, date, timedelta
import json, hashlib, uuid, calendar, io
from collections import OrderedDict
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ─────────────────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────────────────
st.set_page_config(page_title="AM Task Tracker", page_icon="📊", layout="wide")

TASK_COLS     = ["Review Meeting", "PPC Meeting", "Presales Review", "Meeting CP Agg", "MOM Nurturing"]
WEEKLY_TARGET = 5

# ─────────────────────────────────────────────────────────
#  SUPABASE REST API  (uses requests — no supabase package)
# ─────────────────────────────────────────────────────────
def _url(table):
    return f"{st.secrets['SUPABASE_URL']}/rest/v1/{table}"

def _headers(prefer=None):
    h = {
        "apikey":        st.secrets["SUPABASE_KEY"],
        "Authorization": f"Bearer {st.secrets['SUPABASE_KEY']}",
        "Content-Type":  "application/json",
    }
    if prefer:
        h["Prefer"] = prefer
    return h

def db_select(table, filters=None, order=None):
    params = {"select": "*"}
    if filters:
        params.update(filters)
    if order:
        params["order"] = order
    r = requests.get(_url(table), headers=_headers(), params=params)
    return r.json() if r.status_code < 300 else []

def db_insert(table, data):
    r = requests.post(_url(table), headers=_headers("return=representation"),
                      json=data if isinstance(data, list) else [data])
    return r.json()

def db_update(table, match_col, match_val, data):
    params = {match_col: f"eq.{match_val}"}
    r = requests.patch(_url(table), headers=_headers("return=representation"),
                       params=params, json=data)
    return r.json()

def db_delete(table, match_col, match_val):
    params = {match_col: f"eq.{match_val}"}
    requests.delete(_url(table), headers=_headers(), params=params)

def db_upsert(table, data, on_conflict):
    r = requests.post(_url(table),
                      headers={**_headers(), "Prefer": f"resolution=merge-duplicates,return=representation"},
                      params={"on_conflict": on_conflict},
                      json=data if isinstance(data, list) else [data])
    return r.json()

# ─────────────────────────────────────────────────────────
#  DATE HELPERS
# ─────────────────────────────────────────────────────────
def get_week_start(d=None):
    d = d or date.today()
    if isinstance(d, str): d = date.fromisoformat(d)
    return d - timedelta(days=d.weekday())

def week_label(ws):
    if isinstance(ws, str): ws = date.fromisoformat(ws)
    we = ws + timedelta(days=4)
    return f"{ws.strftime('%d')}-{we.strftime('%d %b %Y')}"

def month_label(year, month):
    return f"{calendar.month_name[month]} {year}"

def get_weeks_in_month(year, month):
    first = date(year, month, 1)
    last  = date(year, month, calendar.monthrange(year, month)[1])
    weeks, cur = [], first - timedelta(days=first.weekday())
    while cur <= last:
        if cur.month == month or (cur + timedelta(days=4)).month == month:
            weeks.append(str(cur))
        cur += timedelta(weeks=1)
    return weeks

def hp(pw): return hashlib.sha256(pw.encode()).hexdigest()

# ─────────────────────────────────────────────────────────
#  DATA OPERATIONS
# ─────────────────────────────────────────────────────────
def get_users():
    rows = db_select("users")
    return {r["username"]: r for r in rows}

def get_projects():
    return db_select("projects")

def add_project(name, am1, am2):
    db_insert("projects", {"id": str(uuid.uuid4()), "name": name, "am1": am1, "am2": am2})

def delete_project(pid):
    db_delete("projects", "id", pid)

def get_tasks_for_week(week_start):
    return db_select("tasks", {"week_start": f"eq.{week_start}"})

def get_all_tasks():
    return db_select("tasks", order="week_start.desc")

def upsert_task(project_id, am1, am2, task_col, status, ws):
    # status: "done", "na", "pending"
    ws_str = str(ws)
    now    = str(datetime.now())
    checked = (status == "done")
    existing = db_select("tasks", {
        "project_id": f"eq.{project_id}",
        "week_start": f"eq.{ws_str}",
        "task_col":   f"eq.{task_col}"
    })
    if existing:
        db_update("tasks", "id", existing[0]['id'],
                  {"checked": checked, "status": status, "updated_at": now})
    else:
        db_insert("tasks", {
            "id": str(uuid.uuid4()),
            "project_id": project_id,
            "am1": am1, "am2": am2,
            "task_col": task_col,
            "checked":  checked,
            "status":   status,
            "week_start": ws_str,
            "month_label": month_label(ws.year, ws.month),
            "month": ws.month, "year": ws.year,
            "created_at": now, "updated_at": now,
        })

def get_weekly_reports():
    rows = db_select("weekly_reports", order="week_start.desc")
    for r in rows:
        if isinstance(r.get("rows"), str):
            r["rows"] = json.loads(r["rows"])
    return rows

def save_weekly_report(report):
    payload = {**report, "rows": json.dumps(report["rows"])}
    db_upsert("weekly_reports", payload, "week_start")

def get_monthly_reports():
    rows = db_select("monthly_reports", order="year.desc")
    for r in rows:
        if isinstance(r.get("rows"),  str): r["rows"]  = json.loads(r["rows"])
        if isinstance(r.get("weeks"), str): r["weeks"] = json.loads(r["weeks"])
    return rows

def save_monthly_report(report):
    payload = {**report,
               "rows":  json.dumps(report["rows"]),
               "weeks": json.dumps(report["weeks"])}
    db_upsert("monthly_reports", payload, "year,month")

def seed():
    users = get_users()
    defaults = [
        {"username":"manager",   "name":"Manager",   "role":"manager",   "password":hp("manager123")},
        {"username":"executive", "name":"Executive", "role":"executive", "password":hp("exec123")},
    ]
    for d in defaults:
        if d["username"] not in users:
            db_insert("users", d)

    if not get_projects():
        db_insert("projects", [
            {"id":"p1",  "name":"Sai Vaastu Dreams",  "am1":"Aadam",  "am2":"Dikshant"},
            {"id":"p2",  "name":"Unity Towers",        "am1":"Aadam",  "am2":"Dikshant"},
            {"id":"p3",  "name":"The Landmark",        "am1":"Aadam",  "am2":"Dikshant"},
            {"id":"p4",  "name":"Elate Residences",    "am1":"Aadam",  "am2":"Saquib"},
            {"id":"p5",  "name":"Somerset Park",       "am1":"Aadam",  "am2":"Saquib"},
            {"id":"p6",  "name":"Sukhwani Verde",      "am1":"Aadam",  "am2":"Saquib"},
            {"id":"p7",  "name":"Sukhwani Skylines",   "am1":"Swarni", "am2":"Bhumika"},
            {"id":"p8",  "name":"Aishwaryam Abhimaan", "am1":"Swarni", "am2":"Bhumika"},
            {"id":"p9",  "name":"Treasure Troves",     "am1":"Swarni", "am2":"Anurag L"},
            {"id":"p10", "name":"The Shashwat",        "am1":"Swarni", "am2":"Anurag L"},
            {"id":"p11", "name":"Silverian Mall",      "am1":"Swarni", "am2":"Anurag L"},
        ])

# ─────────────────────────────────────────────────────────
#  REPORT GENERATORS
# ─────────────────────────────────────────────────────────
def generate_weekly_report(week_start_str):
    projects = get_projects()
    wt       = get_tasks_for_week(week_start_str)
    rows = []
    for p in projects:
        task_map = {}
        for t in wt:
            if t["project_id"] == p["id"]:
                task_map[t["task_col"]] = t.get("status", "done" if t.get("status","done" if t.get("checked") else "pending")=="done" else "pending")
        row = {"project_id":p["id"],"project_name":p["name"],"am1":p["am1"],"am2":p["am2"]}
        applicable = 0
        scored     = 0
        for tc in TASK_COLS:
            st_val = task_map.get(tc, "pending")
            row[tc] = st_val
            if st_val != "na":
                applicable += 1
                if st_val == "done":
                    scored += 1
        row["score"]       = scored
        row["applicable"]  = applicable
        row["max_possible"]= applicable
        rows.append(row)
    ws = date.fromisoformat(week_start_str)
    report = {
        "id":          str(uuid.uuid4()),
        "week_start":  week_start_str,
        "week_end":    str(ws + timedelta(days=4)),
        "week_label":  week_label(ws),
        "month":       ws.month, "year": ws.year,
        "month_label": month_label(ws.year, ws.month),
        "generated_at":str(datetime.now()),
        "rows":        rows,
    }
    save_weekly_report(report)
    return report

def generate_monthly_report(year, month):
    projects  = get_projects()
    weeks     = get_weeks_in_month(year, month)
    all_tasks = get_all_tasks()
    mt = [t for t in all_tasks if t["week_start"] in weeks]
    rows = []
    for p in projects:
        row = {"project_id":p["id"],"project_name":p["name"],"am1":p["am1"],"am2":p["am2"]}
        total_scored = 0
        total_applicable = 0
        for tc in TASK_COLS:
            proj_tasks = [t for t in mt if t["project_id"]==p["id"] and t["task_col"]==tc]
            done  = sum(1 for t in proj_tasks if t.get("status","done" if t.get("status","done" if t.get("checked") else "pending")=="done" else "pending")=="done")
            na    = sum(1 for t in proj_tasks if t.get("status","")=="na")
            applicable = len(weeks) - na
            row[tc] = done
            total_scored     += done
            total_applicable += max(applicable, 0)
        row["total_score"]    = total_scored
        row["weeks_tracked"]  = len(weeks)
        row["max_possible"]   = total_applicable
        row["completion_pct"] = round(total_scored/total_applicable*100,1) \
                                if total_applicable else 0
        rows.append(row)
    report = {
        "id":          str(uuid.uuid4()),
        "year": year,  "month": month,
        "month_label": month_label(year, month),
        "weeks":       weeks,
        "generated_at":str(datetime.now()),
        "rows":        rows,
    }
    save_monthly_report(report)
    return report

# ─────────────────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────────────────
_YELLOW = PatternFill("solid", fgColor="FFFF00")
_PINK   = PatternFill("solid", fgColor="FFD7BE")
_BLUE   = PatternFill("solid", fgColor="BDD7EE")
_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_WHITE  = PatternFill("solid", fgColor="FFFFFF")
_RED    = PatternFill("solid", fgColor="FF4D4D")
_AMBER  = PatternFill("solid", fgColor="FFD700")
_LIME   = PatternFill("solid", fgColor="92D050")
_thin   = Side(style="thin", color="000000")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _sfill(s):
    s = int(s)
    return _RED if s==0 else (_AMBER if s<5 else _LIME)

def _c(ws, row, col, value="", bold=False, fill=None, align="left", size=10):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = Font(bold=bold, name="Arial", size=size, color="000000")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border    = _BORDER
    if fill: cell.fill = fill
    return cell

def build_excel_report(report_rows, report_label, weekly=True):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    for col, w in [("A",18),("B",18),("C",30),("D",16),("E",14),
                   ("F",4),("G",20),("H",14),("I",14)]:
        ws.column_dimensions[col].width = w
    ws.row_dimensions[1].height = 22
    score_hdr = "Sum of Score" if weekly else "Total Score"
    right_hdr = "Total Off"    if weekly else "Completion %"
    for ci, h in enumerate(["AM 1","AM 2","Project",score_hdr,right_hdr], 1):
        _c(ws, 1, ci, h, bold=True, fill=_BLUE, align="center")
    ws.freeze_panes = "A2"
    am1_map = OrderedDict()
    for r in report_rows:
        am1_map.setdefault(r["am1"], OrderedDict()).setdefault(r["am2"], []).append(r)
    cur_row = 2
    am2_summary = []
    for am1, am2_dict in am1_map.items():
        am1_total = sum(r.get("score",0) if weekly else r.get("total_score",0)
                        for rows in am2_dict.values() for r in rows)
        for ci in range(1, 6):
            val  = am1 if ci==1 else (am1_total if ci==4 else "")
            _c(ws, cur_row, ci, val, bold=ci in (1,4), fill=_YELLOW,
               align="center" if ci in (1,4) else "left", size=11)
        cur_row += 1
        for am2, projs in am2_dict.items():
            am2_total = sum(r.get("score",0) if weekly else r.get("total_score",0)
                            for r in projs)
            max_pts   = len(projs) * WEEKLY_TARGET
            right_val = (f"{am2_total} off {max_pts}" if weekly
                         else f"{round(am2_total/max_pts*100,0):.0f}%" if max_pts else "0%")
            _c(ws, cur_row, 1, "",        fill=_PINK)
            _c(ws, cur_row, 2, am2,       bold=True, fill=_PINK)
            _c(ws, cur_row, 3, "",        fill=_PINK)
            _c(ws, cur_row, 4, am2_total, bold=True, fill=_PINK, align="center")
            _c(ws, cur_row, 5, right_val, bold=True, fill=_PINK, align="center")
            cur_row += 1
            am2_summary.append({"am2":am2,"obtained":am2_total,"total":max_pts})
            for p in projs:
                sc = p.get("score",0) if weekly else p.get("total_score",0)
                _c(ws, cur_row, 1, "", fill=_WHITE)
                _c(ws, cur_row, 2, "", fill=_WHITE)
                _c(ws, cur_row, 3, p["project_name"], fill=_WHITE)
                _c(ws, cur_row, 4, sc, fill=_sfill(sc), align="center", bold=(sc>=5))
                _c(ws, cur_row, 5, WEEKLY_TARGET if weekly else p.get("max_possible",""),
                   fill=_WHITE, align="center")
                cur_row += 1
    _c(ws, 1, 7, report_label, bold=True, fill=_GREEN, align="center", size=11)
    _c(ws, 1, 8, "", fill=_GREEN)
    _c(ws, 1, 9, "", fill=_GREEN)
    ws.merge_cells("G1:I1")
    for ci, h in enumerate(["AM","Obtained","Total"], 7):
        _c(ws, 2, ci, h, bold=True, fill=_GREEN, align="center")
    for i, s in enumerate(am2_summary, 3):
        _c(ws, i, 7, s["am2"],      bold=s["obtained"]>=5, fill=_WHITE)
        _c(ws, i, 8, s["obtained"], bold=s["obtained"]>=5,
           fill=_sfill(s["obtained"]), align="center")
        _c(ws, i, 9, s["total"],    fill=_WHITE, align="center")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ─────────────────────────────────────────────────────────
#  CSS
# ─────────────────────────────────────────────────────────
def css():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    html,body,[class*="css"]{font-family:'Inter',sans-serif;}
    [data-testid="stSidebar"]{background:linear-gradient(180deg,#0f172a,#1e3a5f);}
    [data-testid="stSidebar"] label,
    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span{color:#e2e8f0 !important;}
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2{color:#f8fafc !important;}
    .ph{background:linear-gradient(135deg,#0f172a,#1e3a5f);color:white;
        padding:1.25rem 1.75rem;border-radius:14px;margin-bottom:1.25rem;}
    .ph h1{margin:0;font-size:1.5rem;font-weight:700;}
    .ph p{margin:.2rem 0 0;opacity:.7;font-size:.85rem;}
    .card{background:white;border:1px solid #e2e8f0;border-radius:12px;
          padding:1rem 1.25rem;margin-bottom:.6rem;box-shadow:0 1px 3px rgba(0,0,0,.06);}
    .filter-bar{background:#f8fafc;border:1px solid #e2e8f0;border-radius:10px;
                padding:.75rem 1rem;margin-bottom:1rem;}
    .score-0   {background:#fff1f2;color:#dc2626;font-weight:800;padding:3px 12px;
                border-radius:6px;display:inline-block;min-width:36px;text-align:center;}
    .score-low {background:#fefce8;color:#b45309;font-weight:800;padding:3px 12px;
                border-radius:6px;display:inline-block;min-width:36px;text-align:center;}
    .score-high{background:#f0fdf4;color:#166534;font-weight:800;padding:3px 12px;
                border-radius:6px;display:inline-block;min-width:36px;text-align:center;}
    [data-testid="metric-container"]{background:#f8fafc;border:1px solid #e2e8f0;
        border-radius:10px;padding:.75rem;}
    .row-text{font-size:.82rem;margin:0;padding:2px 0;}
    .row-bold{font-size:.82rem;margin:0;font-weight:600;padding:2px 0;}
    </style>
    """, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────
#  AUTH
# ─────────────────────────────────────────────────────────
def login():
    st.markdown("""
    <div style="max-width:400px;margin:5rem auto;background:white;border-radius:20px;
                padding:2.5rem;box-shadow:0 8px 32px rgba(0,0,0,.12);border:1px solid #e2e8f0;">
        <div style="text-align:center;margin-bottom:1.75rem">
            <div style="font-size:2.5rem">📊</div>
            <h2 style="margin:.4rem 0 .2rem;color:#0f172a;font-weight:700">AM Task Tracker</h2>
            <p style="color:#64748b;margin:0;font-size:.88rem">Account Manager Performance System</p>
        </div>
    </div>""", unsafe_allow_html=True)
    c1,c2,c3 = st.columns([1,2,1])
    with c2:
        with st.form("lf"):
            un = st.text_input("Username")
            pw = st.text_input("Password", type="password")
            ok = st.form_submit_button("Sign In", use_container_width=True, type="primary")
        if ok:
            users = get_users()
            if un in users and users[un]["password"] == hp(pw):
                st.session_state.update(logged_in=True, username=un,
                    role=users[un]["role"], name=users[un]["name"])
                st.rerun()
            else:
                st.error("Invalid username or password")
        st.markdown("""
        <div style="margin-top:1rem;background:#f1f5f9;border-radius:8px;padding:.75rem;
                    font-size:.78rem;color:#475569">
        <b>Login:</b><br>
        Manager &nbsp;&nbsp;: <code>manager</code> / <code>manager123</code><br>
        Executive : <code>executive</code> / <code>exec123</code>
        </div>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────────────────
def sidebar(role):
    rc = "#3b82f6" if role=="manager" else "#8b5cf6"
    with st.sidebar:
        st.markdown(f"""
        <div style="padding:.75rem 0 1.25rem;text-align:center">
            <div style="font-size:2rem">📊</div>
            <h2 style="margin:.4rem 0 .2rem;font-size:1rem;font-weight:700">AM Task Tracker</h2>
        </div>
        <div style="background:rgba(255,255,255,.08);border-radius:10px;
                    padding:.65rem 1rem;margin-bottom:1.25rem">
            <p style="margin:0;font-size:.72rem;opacity:.55">Logged in as</p>
            <p style="margin:0;font-weight:600;font-size:.95rem">{st.session_state['name']}</p>
            <span style="font-size:.7rem;background:{rc};padding:2px 8px;
                  border-radius:999px;color:white">{role.upper()}</span>
        </div>""", unsafe_allow_html=True)
        if role=="manager":
            pages = ["🏠 Dashboard","📅 Weekly Report","📆 Monthly Report",
                     "🗂️ Report History","📋 All Tasks","⚙️ Manage"]
        else:
            pages = ["🏠 Overview","✅ Mark Tasks","📅 Weekly Report","📆 Monthly Report"]
        page = st.radio("Nav", pages, label_visibility="collapsed")
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🚪 Logout", use_container_width=True):
            for k in list(st.session_state.keys()): del st.session_state[k]
            st.rerun()
    return page

# ─────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────
def score_html(s):
    s   = int(s)
    cls = "score-0" if s==0 else ("score-low" if s<WEEKLY_TARGET else "score-high")
    return f'<span class="{cls}">{s}</span>'

def color_score_cell(val):
    if not isinstance(val,(int,float)): return ""
    if val==0:  return "color:#dc2626;font-weight:800;background-color:#fff1f2;text-align:center;"
    elif val<5: return "color:#b45309;font-weight:800;background-color:#fefce8;text-align:center;"
    else:       return "color:#166534;font-weight:800;background-color:#f0fdf4;text-align:center;"

def color_task_cell(val):
    if val == "done" or val == 1:  return "color:#166534;font-weight:600;background-color:#f0fdf4;"
    if val == "na":                return "color:#6366f1;font-weight:600;background-color:#eef2ff;"
    if val == "pending" or val == 0: return "color:#94a3b8;background-color:#f8fafc;"
    return ""

def color_pct_cell(val):
    if not isinstance(val,(int,float)): return ""
    if val==0:   return "color:#dc2626;font-weight:800;background-color:#fff1f2;text-align:center;"
    elif val<60: return "color:#b45309;font-weight:800;background-color:#fefce8;text-align:center;"
    else:        return "color:#166534;font-weight:800;background-color:#f0fdf4;text-align:center;"

# ─────────────────────────────────────────────────────────
#  FILTER BAR
# ─────────────────────────────────────────────────────────
def filter_bar(df, show_week=True, show_month=True, show_am1=True,
               show_am2=True, show_project=True, key_prefix="fb"):
    st.markdown('<div class="filter-bar">', unsafe_allow_html=True)
    st.markdown("**🔍 Filters**")
    proj_col = "project_name" if "project_name" in df.columns else \
               "Project"      if "Project"       in df.columns else None
    num = sum([show_month, show_week, show_am1, show_am2,
               bool(show_project and proj_col)])
    cols = st.columns(num) if num else []
    ci = 0
    f_month=f_week=f_am1=f_am2=f_proj = "All"

    if show_month and "month_label" in df.columns:
        months  = ["All"]+sorted(df["month_label"].dropna().unique().tolist(), reverse=True)
        f_month = cols[ci].selectbox("📆 Month", months, key=f"{key_prefix}_month"); ci+=1
    if show_week and "week_start" in df.columns:
        wkopts = ["All"]+sorted(df["week_start"].dropna().unique().tolist(), reverse=True)
        f_week = cols[ci].selectbox("📅 Week", wkopts,
                    format_func=lambda w: week_label(w) if w!="All" else "All",
                    key=f"{key_prefix}_week"); ci+=1
    if show_am1 and "am1" in df.columns:
        am1s  = ["All"]+sorted(df["am1"].dropna().unique().tolist())
        f_am1 = cols[ci].selectbox("👤 AM1", am1s, key=f"{key_prefix}_am1"); ci+=1
    if show_am2 and "am2" in df.columns:
        sub   = df if f_am1=="All" else df[df["am1"]==f_am1]
        am2s  = ["All"]+sorted(sub["am2"].dropna().unique().tolist())
        f_am2 = cols[ci].selectbox("👥 AM2", am2s, key=f"{key_prefix}_am2"); ci+=1
    if show_project and proj_col:
        sub2 = df.copy()
        if f_am1!="All": sub2=sub2[sub2["am1"]==f_am1]
        if f_am2!="All": sub2=sub2[sub2["am2"]==f_am2]
        projs  = ["All"]+sorted(sub2[proj_col].dropna().unique().tolist())
        f_proj = cols[ci].selectbox("🏗️ Project", projs, key=f"{key_prefix}_proj")

    st.markdown('</div>', unsafe_allow_html=True)
    fdf = df.copy()
    if f_month!="All" and "month_label" in fdf.columns: fdf=fdf[fdf["month_label"]==f_month]
    if f_week !="All" and "week_start"  in fdf.columns: fdf=fdf[fdf["week_start"] ==f_week]
    if f_am1  !="All" and "am1" in fdf.columns:         fdf=fdf[fdf["am1"]==f_am1]
    if f_am2  !="All" and "am2" in fdf.columns:         fdf=fdf[fdf["am2"]==f_am2]
    if f_proj !="All" and proj_col:                      fdf=fdf[fdf[proj_col]==f_proj]
    return fdf, {"month":f_month,"week":f_week,"am1":f_am1,"am2":f_am2,"project":f_proj}

# ─────────────────────────────────────────────────────────
#  SHARED REPORT RENDERERS
# ─────────────────────────────────────────────────────────
def render_weekly_report(report, key_prefix="wr"):
    rows = report.get("rows",[])
    if not rows: st.info("No data."); return
    df = pd.DataFrame(rows)
    df["week_start"]  = report["week_start"]
    df["month_label"] = report.get("month_label","")

    # Normalise task columns — old data has 0/1, new data has "done"/"na"/"pending"
    for tc in TASK_COLS:
        if tc in df.columns:
            df[tc] = df[tc].apply(lambda v:
                "done" if v in (1, True, "done") else
                "na"   if v == "na" else "pending")

    # Ensure score & applicable columns exist
    if "score" not in df.columns:
        df["score"] = df.apply(lambda r: sum(1 for tc in TASK_COLS if r.get(tc)=="done"), axis=1)
    df["score"] = pd.to_numeric(df["score"], errors="coerce").fillna(0).astype(int)
    if "applicable" not in df.columns:
        df["applicable"] = df.apply(lambda r: sum(1 for tc in TASK_COLS if r.get(tc)!="na"), axis=1)

    st.markdown(f"#### 📅 {report['week_label']}")
    st.caption(f"Generated: {report['generated_at'][:19]}")
    fdf, sel = filter_bar(df, show_week=False, show_month=False,
                          show_am1=True, show_am2=True, show_project=True,
                          key_prefix=key_prefix)
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Total Points",        int(fdf["score"].sum()))
    c2.metric("Projects Hit Target", int((fdf["score"] >= fdf["applicable"]).sum()))
    c3.metric("Zero-Score",          int((fdf["score"]==0).sum()))
    c4.metric("Avg Score",           round(fdf["score"].mean(),1) if len(fdf) else 0)

    st.markdown("##### Pivot Table")
    disp = fdf[["am1","am2","project_name"]+TASK_COLS+["score","applicable"]].copy()
    disp["Score"] = disp.apply(lambda r: f"{r['score']}/{r['applicable']}", axis=1)
    disp = disp[["am1","am2","project_name"]+TASK_COLS+["Score"]]
    disp.columns = ["AM1","AM2","Project"]+TASK_COLS+["Score"]
    styled = (disp.style
              .map(color_task_cell,  subset=TASK_COLS)
              .set_properties(**{"text-align":"center"}, subset=TASK_COLS+["Score"]))
    st.dataframe(styled, use_container_width=True, hide_index=True)

    if sel["project"]=="All":
        st.markdown("##### AM2 Summary")
        am2s = fdf.groupby(["am1","am2"]).agg(
            Projects=("project_name","count"),
            Total=("score","sum"), Avg=("score","mean")).reset_index()
        am2s.columns=["AM1","AM2","Projects","Total Points","Avg Score"]
        am2s["Avg Score"]=am2s["Avg Score"].round(1)
        st.dataframe(am2s, use_container_width=True, hide_index=True)
        fig = go.Figure()
        fig.add_bar(x=am2s["AM2"], y=am2s["Total Points"],
                    marker_color=["#22c55e" if s>=10 else "#f59e0b" if s>0 else "#ef4444"
                                  for s in am2s["Total Points"]],
                    text=am2s["Total Points"], textposition="outside")
        fig.update_layout(title="Points by AM2", template="plotly_white",
                          showlegend=False, margin=dict(t=40,b=30,l=20,r=20))
        st.plotly_chart(fig, use_container_width=True)
    c1,c2 = st.columns(2)
    c1.download_button("📥 CSV", disp.to_csv(index=False).encode(),
                       f"weekly_{report['week_start']}.csv","text/csv",
                       key=f"dl_{key_prefix}")
    xlsx = build_excel_report(fdf.to_dict("records"), report["week_label"], weekly=True)
    c2.download_button("📊 Excel", xlsx,
                       f"weekly_{report['week_start']}.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       key=f"xl_{key_prefix}")


def render_monthly_report(report, key_prefix="mr"):
    rows = report.get("rows",[])
    if not rows: st.info("No data."); return
    df = pd.DataFrame(rows)
    df["month_label"] = report["month_label"]

    # Ensure numeric columns
    for col in ["total_score","completion_pct","max_possible"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # Normalise task columns to numeric done-count
    for tc in TASK_COLS:
        if tc in df.columns:
            df[tc] = pd.to_numeric(df[tc], errors="coerce").fillna(0).astype(int)

    st.markdown(f"#### 📆 {report['month_label']}")
    st.caption(f"Weeks: {len(report['weeks'])} | Generated: {report['generated_at'][:19]}")
    fdf, sel = filter_bar(df, show_week=False, show_month=False,
                          show_am1=True, show_am2=True, show_project=True,
                          key_prefix=key_prefix)
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Total Points",        int(fdf["total_score"].sum()))
    c2.metric("Projects Hit Target", int((fdf["completion_pct"]>=100).sum()))
    c3.metric("Zero-Score",          int((fdf["total_score"]==0).sum()))
    c4.metric("Avg Completion",
              f"{round(fdf['completion_pct'].mean(),1)}%" if len(fdf) else "0%")
    st.markdown("##### Monthly Pivot Table")
    disp = fdf[["am1","am2","project_name"]+TASK_COLS+["total_score","completion_pct"]].copy()
    disp.columns = ["AM1","AM2","Project"]+TASK_COLS+["Total Score","Completion %"]
    styled = (disp.style
              .map(color_score_cell, subset=["Total Score"])
              .map(color_pct_cell,   subset=["Completion %"])
              .map(color_task_cell,  subset=TASK_COLS)
              .set_properties(**{"text-align":"center"},
                              subset=TASK_COLS+["Total Score","Completion %"]))
    st.dataframe(styled, use_container_width=True, hide_index=True)
    if sel["project"]=="All":
        st.markdown("##### AM2 Monthly Summary")
        am2s = fdf.groupby(["am1","am2"]).agg(
            Projects=("project_name","count"),
            Total=("total_score","sum"),
            Avg_Pct=("completion_pct","mean")).reset_index()
        am2s.columns=["AM1","AM2","Projects","Total Score","Avg Completion %"]
        am2s["Avg Completion %"]=am2s["Avg Completion %"].round(1)
        st.dataframe(am2s, use_container_width=True, hide_index=True)
        task_totals = {tc:int(fdf[tc].sum()) for tc in TASK_COLS}
        fig = go.Figure()
        fig.add_bar(x=list(task_totals.keys()), y=list(task_totals.values()),
                    marker_color="#3b82f6", text=list(task_totals.values()),
                    textposition="outside")
        fig.update_layout(title="Completions per Task Type", template="plotly_white",
                          showlegend=False, margin=dict(t=40,b=30,l=20,r=20))
        st.plotly_chart(fig, use_container_width=True)
        am2t = fdf.groupby("am2")[TASK_COLS].sum().reset_index()
        fig2 = px.bar(am2t.melt(id_vars="am2",var_name="Task",value_name="Count"),
                      x="am2",y="Count",color="Task",barmode="group",
                      title="Task Breakdown by AM2",
                      color_discrete_sequence=px.colors.qualitative.Set2)
        fig2.update_layout(template="plotly_white",margin=dict(t=40,b=30,l=20,r=20))
        st.plotly_chart(fig2, use_container_width=True)
    c1,c2 = st.columns(2)
    c1.download_button("📥 CSV", disp.to_csv(index=False).encode(),
                       f"monthly_{report['year']}_{report['month']:02d}.csv","text/csv",
                       key=f"dl_{key_prefix}")
    xlsx = build_excel_report(fdf.to_dict("records"), report["month_label"], weekly=False)
    c2.download_button("📊 Excel", xlsx,
                       f"monthly_{report['year']}_{report['month']:02d}.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       key=f"xl_{key_prefix}")

# ─────────────────────────────────────────────────────────
#  EXECUTIVE PAGES
# ─────────────────────────────────────────────────────────
def exec_overview():
    st.markdown(f'<div class="ph"><h1>🏠 Overview</h1>'
                f'<p>{week_label(get_week_start())}</p></div>', unsafe_allow_html=True)
    ws       = str(get_week_start())
    projects = get_projects()
    wt       = get_tasks_for_week(ws)
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Total Projects",    len(projects))
    c2.metric("Active This Week",  len({t["project_id"] for t in wt if t.get("status","done" if t.get("checked") else "pending")=="done"}))
    c3.metric("Tasks Checked Off", sum(1 for t in wt if t.get("status","done" if t.get("checked") else "pending")=="done"))
    c4.metric("Zero-Score",
              sum(1 for p in projects
                  if not any(t["project_id"]==p["id"] and t.get("status","done" if t.get("checked") else "pending")=="done" for t in wt)))
    rows = [{"am1":p["am1"],"am2":p["am2"],"project_name":p["name"],
             "score":sum(1 for t in wt if t["project_id"]==p["id"] and t.get("status","done" if t.get("checked") else "pending")=="done")}
            for p in projects]
    if rows:
        df = pd.DataFrame(rows)
        fdf,_ = filter_bar(df, show_week=False, show_month=False,
                           show_am1=True, show_am2=True, show_project=True,
                           key_prefix="ov_exec")
        disp = fdf[["am1","am2","project_name","score"]].copy()
        disp.columns = ["AM1","AM2","Project","Score"]
        st.dataframe(disp.style.map(color_score_cell, subset=["Score"]),
                     use_container_width=True, hide_index=True)


def exec_mark_tasks():
    # ── Week selector ──
    today        = date.today()
    week_options = [str(get_week_start(today - timedelta(weeks=i))) for i in range(8)]
    selected_ws  = st.selectbox("📅 Select Week", week_options,
                                format_func=lambda w: week_label(w) + (" · Current" if w==week_options[0] else ""),
                                key="mt_week")
    ws = date.fromisoformat(selected_ws)

    st.markdown(f'<div class="ph"><h1>✅ Mark Weekly Tasks</h1>'
                f'<p>{week_label(ws)} · Tick done · Check NA box if task not applicable</p></div>',
                unsafe_allow_html=True)

    projects = get_projects()
    if not projects: st.info("No projects yet."); return
    wt = get_tasks_for_week(str(ws))

    # ── Filters ──
    all_am1 = sorted({p["am1"] for p in projects})
    c1,c2,c3 = st.columns(3)
    f_am1 = c1.selectbox("Filter AM1", ["All"]+all_am1, key="mt_am1")
    f_am2 = c2.selectbox("Filter AM2",
                         ["All"]+sorted({p["am2"] for p in projects
                                         if f_am1=="All" or p["am1"]==f_am1}), key="mt_am2")
    f_proj_names = sorted({p["name"] for p in projects
                           if (f_am1=="All" or p["am1"]==f_am1)
                           and (f_am2=="All" or p["am2"]==f_am2)})
    f_proj = c3.selectbox("Filter Project", ["All"]+f_proj_names, key="mt_proj")
    filtered = [p for p in projects
                if (f_am1=="All" or p["am1"]==f_am1)
                and (f_am2=="All" or p["am2"]==f_am2)
                and (f_proj=="All" or p["name"]==f_proj)]

    # ── Column layout ──
    COLS = [0.25, 1.1, 1.1, 2.2, 0.85, 0.85, 0.95, 0.85, 0.85, 0.6]
    HDR  = ["Sr.", "AM1", "AM2", "Project",
            "Review\nMtg", "PPC\nMtg", "Presales\nRev", "Mtg\nCP Agg", "MOM\nNurt.", "Score"]
    hrow = st.columns(COLS)
    for col, h in zip(hrow, HDR):
        col.markdown(
            f"<div style='background:#1e3a5f;color:white;font-weight:700;font-size:.68rem;"
            f"text-align:center;padding:5px 2px;border-radius:6px;line-height:1.3;"
            f"min-height:34px;display:flex;align-items:center;justify-content:center'>{h}</div>",
            unsafe_allow_html=True)
    st.divider()

    sr = 1
    for p in filtered:
        # Build status map
        task_status = {}
        for t in wt:
            if t["project_id"] == p["id"]:
                task_status[t["task_col"]] = t.get("status",
                    "done" if t.get("status","done" if t.get("checked") else "pending")=="done" else "pending")

        applicable = sum(1 for tc in TASK_COLS if task_status.get(tc, "pending") != "na")
        scored     = sum(1 for tc in TASK_COLS if task_status.get(tc, "pending") == "done")

        # ── Done checkboxes row ──
        row = st.columns(COLS)
        row[0].markdown(f"<p class='row-text' style='text-align:center;padding-top:6px'>{sr}</p>",
                        unsafe_allow_html=True)
        row[1].markdown(f"<p class='row-text' style='padding-top:6px'>{p['am1']}</p>",
                        unsafe_allow_html=True)
        row[2].markdown(f"<p class='row-text' style='padding-top:6px'>{p['am2']}</p>",
                        unsafe_allow_html=True)
        row[3].markdown(f"<p class='row-bold' style='padding-top:6px'>{p['name']}</p>",
                        unsafe_allow_html=True)

        for i, tc in enumerate(TASK_COLS):
            cur   = task_status.get(tc, "pending")
            is_na = (cur == "na")
            if is_na:
                row[4+i].markdown(
                    "<div style='text-align:center;color:#818cf8;font-size:.75rem;"
                    "font-weight:700;padding-top:8px'>N/A</div>",
                    unsafe_allow_html=True)
            else:
                new_done = row[4+i].checkbox("Done", value=(cur=="done"),
                                             key=f"done_{p['id']}_{tc}_{selected_ws}")
                if new_done != (cur=="done"):
                    upsert_task(p["id"], p["am1"], p["am2"], tc,
                                "done" if new_done else "pending", ws)
                    st.rerun()

        # Score
        sc_class = ("score-high" if scored==applicable and applicable>0
                    else "score-0" if scored==0
                    else "score-low")
        row[9].markdown(
            f"<div style='padding-top:6px;text-align:center'>"
            f"<span class='{sc_class}'>{scored}/{applicable}</span></div>",
            unsafe_allow_html=True)

        # ── NA checkboxes row (small, subtle) ──
        na_row = st.columns(COLS)
        na_row[3].markdown(
            "<p style='font-size:.62rem;color:#94a3b8;text-align:right;"
            "margin:0;padding-right:6px;padding-top:2px'>N/A →</p>",
            unsafe_allow_html=True)
        for i, tc in enumerate(TASK_COLS):
            cur   = task_status.get(tc, "pending")
            is_na = (cur == "na")
            new_na = na_row[4+i].checkbox("N/A", value=is_na,
                                          key=f"na_{p['id']}_{tc}_{selected_ws}")
            if new_na != is_na:
                upsert_task(p["id"], p["am1"], p["am2"], tc,
                            "na" if new_na else "pending", ws)
                st.rerun()

        st.markdown("<hr style='margin:.25rem 0;border-color:#f1f5f9'>",
                    unsafe_allow_html=True)
        sr += 1

    st.markdown(f"<p style='color:#94a3b8;font-size:.78rem;margin-top:.5rem'>"
                f"Showing {sr-1} of {len(projects)} projects · Changes auto-saved · "
                f"Score = Done / Applicable (N/A excluded)</p>",
                unsafe_allow_html=True)




def exec_weekly_page():
    st.markdown('<div class="ph"><h1>📅 Weekly Report</h1></div>', unsafe_allow_html=True)
    all_tasks = get_all_tasks()
    if not all_tasks: st.info("No data yet. Mark some tasks first."); return
    all_weeks = sorted({t["week_start"] for t in all_tasks}, reverse=True)
    week_sel  = st.selectbox("Select Week", all_weeks, format_func=week_label)
    c1,c2 = st.columns([3,1])
    with c2:
        if st.button("🔄 Generate / Refresh", type="primary", use_container_width=True):
            with st.spinner("Generating..."): generate_weekly_report(week_sel)
            st.success("Saved!"); st.rerun()
    reports = get_weekly_reports()
    saved = next((r for r in reports if r["week_start"]==week_sel), None)
    if saved: render_weekly_report(saved, key_prefix="exec_wr")
    else:     st.info("Click **Generate / Refresh** to create this report.")


def exec_monthly_page():
    st.markdown('<div class="ph"><h1>📆 Monthly Report</h1></div>', unsafe_allow_html=True)
    today = date.today()
    opts  = [(date(today.year,today.month,1)-timedelta(days=30*i)) for i in range(12)]
    opts  = [(d.year,d.month) for d in opts]
    labels= [month_label(y,m) for y,m in opts]
    sel   = st.selectbox("Select Month", labels)
    sy,sm = opts[labels.index(sel)]
    c1,c2 = st.columns([3,1])
    with c2:
        if st.button("🔄 Generate / Refresh", type="primary", use_container_width=True):
            with st.spinner("Generating..."): generate_monthly_report(sy,sm)
            st.success("Saved!"); st.rerun()
    reports = get_monthly_reports()
    saved = next((r for r in reports if r["year"]==sy and r["month"]==sm), None)
    if saved: render_monthly_report(saved, key_prefix="exec_mr")
    else:     st.info("Click **Generate / Refresh** to create this report.")

# ─────────────────────────────────────────────────────────
#  MANAGER PAGES
# ─────────────────────────────────────────────────────────
def mgr_dashboard():
    st.markdown(f'<div class="ph"><h1>🏠 Manager Dashboard</h1>'
                f'<p>{week_label(get_week_start())}</p></div>', unsafe_allow_html=True)
    ws       = str(get_week_start())
    projects = get_projects()
    wt       = get_tasks_for_week(ws)
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Total Projects",   len(projects))
    c2.metric("Active This Week", len({t["project_id"] for t in wt if t.get("status","done" if t.get("checked") else "pending")=="done"}))
    c3.metric("Total Points",     sum(1 for t in wt if t.get("status","done" if t.get("checked") else "pending")=="done"))
    c4.metric("Zero-Score",
              sum(1 for p in projects
                  if not any(t["project_id"]==p["id"] and t.get("status","done" if t.get("checked") else "pending")=="done" for t in wt)))
    rows = [{"am1":p["am1"],"am2":p["am2"],"project_name":p["name"],
             "score":sum(1 for t in wt if t["project_id"]==p["id"] and t.get("status","done" if t.get("checked") else "pending")=="done")}
            for p in projects]
    if not rows: return
    df  = pd.DataFrame(rows)
    fdf,sel = filter_bar(df, show_week=False, show_month=False,
                         show_am1=True, show_am2=True, show_project=True,
                         key_prefix="dash_mgr")
    if sel["project"]=="All":
        st.markdown("### AM2-wise Progress")
        for am1 in sorted(fdf["am1"].unique()):
            sub = fdf[fdf["am1"]==am1]
            st.markdown(f"**🟡 {am1}**")
            for am2 in sorted(sub["am2"].unique()):
                sub2    = sub[sub["am2"]==am2]
                total   = int(sub2["score"].sum())
                max_pts = len(sub2)*WEEKLY_TARGET
                pct     = min(int(total/max_pts*100),100) if max_pts else 0
                color   = "#22c55e" if pct>=80 else "#f59e0b" if pct>=40 else "#ef4444"
                st.markdown(f"""
                <div class="card">
                  <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:.4rem">
                    <p style="margin:0;font-weight:600;font-size:.9rem">{am2}</p>
                    <span style="font-size:.84rem;color:#64748b">{total}/{max_pts} pts · {len(sub2)} projects</span>
                  </div>
                  <div style="background:#e2e8f0;border-radius:999px;height:8px;overflow:hidden">
                    <div style="width:{pct}%;height:8px;background:{color};border-radius:999px"></div>
                  </div>
                </div>""", unsafe_allow_html=True)


def mgr_weekly_report():
    st.markdown('<div class="ph"><h1>📅 Weekly Report</h1></div>', unsafe_allow_html=True)
    all_tasks = get_all_tasks()
    if not all_tasks: st.info("No data yet."); return
    all_weeks = sorted({t["week_start"] for t in all_tasks}, reverse=True)
    week_sel  = st.selectbox("Select Week", all_weeks, format_func=week_label)
    c1,c2 = st.columns([3,1])
    with c2:
        if st.button("🔄 Generate / Refresh", type="primary", use_container_width=True):
            with st.spinner("Generating..."): generate_weekly_report(week_sel)
            st.success("Saved!"); st.rerun()
    reports = get_weekly_reports()
    saved = next((r for r in reports if r["week_start"]==week_sel), None)
    if saved: render_weekly_report(saved, key_prefix="mgr_wr")
    else:     st.info("Click **Generate / Refresh** to create this report.")


def mgr_monthly_report():
    st.markdown('<div class="ph"><h1>📆 Monthly Report</h1></div>', unsafe_allow_html=True)
    today = date.today()
    opts  = [(date(today.year,today.month,1)-timedelta(days=30*i)) for i in range(24)]
    opts  = [(d.year,d.month) for d in opts]
    labels= [month_label(y,m) for y,m in opts]
    sel   = st.selectbox("Select Month", labels)
    sy,sm = opts[labels.index(sel)]
    c1,c2 = st.columns([3,1])
    with c2:
        if st.button("🔄 Generate / Refresh", type="primary", use_container_width=True):
            with st.spinner("Generating..."): generate_monthly_report(sy,sm)
            st.success("Saved!"); st.rerun()
    reports = get_monthly_reports()
    saved = next((r for r in reports if r["year"]==sy and r["month"]==sm), None)
    if saved: render_monthly_report(saved, key_prefix="mgr_mr")
    else:     st.info("Click **Generate / Refresh** to create this report.")


def mgr_report_history():
    st.markdown('<div class="ph"><h1>🗂️ Report History</h1>'
                '<p>Browse any saved past report</p></div>', unsafe_allow_html=True)
    tab1,tab2 = st.tabs(["📅 Weekly History","📆 Monthly History"])
    with tab1:
        reports = get_weekly_reports()
        if not reports: st.info("No weekly reports saved yet.")
        else:
            all_rows=[]
            for r in reports:
                for row in r["rows"]:
                    row["week_start"]     = r["week_start"]
                    row["month_label"]    = r.get("month_label","")
                    row["week_label_str"] = r["week_label"]
                    all_rows.append(row)
            master = pd.DataFrame(all_rows)
            fdf,_ = filter_bar(master, show_week=True, show_month=True,
                               show_am1=True, show_am2=True, show_project=True,
                               key_prefix="wh")
            if fdf.empty: st.warning("No data matches filters.")
            else:
                for ws in sorted(fdf["week_start"].unique(), reverse=True):
                    sub = fdf[fdf["week_start"]==ws]
                    wl  = sub["week_label_str"].iloc[0]
                    with st.expander(f"📅 {wl} — {len(sub)} projects | {int(sub['score'].sum())} pts"):
                        disp = sub[["am1","am2","project_name"]+TASK_COLS+["score"]].copy()
                        disp.columns = ["AM1","AM2","Project"]+TASK_COLS+["Score"]
                        styled=(disp.style
                                .map(color_score_cell, subset=["Score"])
                                .map(color_task_cell,  subset=TASK_COLS)
                                .set_properties(**{"text-align":"center"},subset=TASK_COLS+["Score"]))
                        st.dataframe(styled, use_container_width=True, hide_index=True)
                        c1,c2=st.columns(2)
                        c1.download_button("📥 CSV", disp.to_csv(index=False).encode(),
                                           f"week_{ws}.csv","text/csv",key=f"hist_wc_{ws}")
                        xlsx=build_excel_report(sub.to_dict("records"),wl,weekly=True)
                        c2.download_button("📊 Excel",xlsx,f"week_{ws}.xlsx",
                                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                           key=f"hist_wx_{ws}")
    with tab2:
        reports = get_monthly_reports()
        if not reports: st.info("No monthly reports saved yet.")
        else:
            all_rows=[]
            for r in reports:
                for row in r["rows"]:
                    row["month_label"]=r["month_label"]
                    all_rows.append(row)
            master=pd.DataFrame(all_rows)
            fdf,_=filter_bar(master, show_week=False, show_month=True,
                             show_am1=True, show_am2=True, show_project=True, key_prefix="mh")
            if fdf.empty: st.warning("No data matches filters.")
            else:
                for ml in sorted(fdf["month_label"].unique(), reverse=True):
                    sub=fdf[fdf["month_label"]==ml]
                    with st.expander(f"📆 {ml} — {len(sub)} projects | "
                                     f"{int(sub['total_score'].sum())} pts | "
                                     f"{round(sub['completion_pct'].mean(),1)}% avg"):
                        disp=sub[["am1","am2","project_name"]+TASK_COLS+
                                 ["total_score","completion_pct"]].copy()
                        disp.columns=["AM1","AM2","Project"]+TASK_COLS+["Total","Completion %"]
                        styled=(disp.style
                                .map(color_score_cell,subset=["Total"])
                                .map(color_pct_cell,  subset=["Completion %"])
                                .map(color_task_cell, subset=TASK_COLS)
                                .set_properties(**{"text-align":"center"},
                                                subset=TASK_COLS+["Total","Completion %"]))
                        st.dataframe(styled, use_container_width=True, hide_index=True)
                        mk=ml.replace(" ","_")
                        c1,c2=st.columns(2)
                        c1.download_button("📥 CSV",disp.to_csv(index=False).encode(),
                                           f"month_{mk}.csv","text/csv",key=f"hist_mc_{mk}")
                        xlsx=build_excel_report(sub.to_dict("records"),ml,weekly=False)
                        c2.download_button("📊 Excel",xlsx,f"month_{mk}.xlsx",
                                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                           key=f"hist_mx_{mk}")


def mgr_all_tasks():
    st.markdown('<div class="ph"><h1>📋 All Tasks</h1>'
                '<p>Complete historical activity log</p></div>', unsafe_allow_html=True)
    all_tasks = get_all_tasks()
    projects  = get_projects()
    if not all_tasks: st.info("No tasks yet."); return
    proj_map = {p["id"]:p for p in projects}
    rows=[]
    for t in all_tasks:
        p=proj_map.get(t["project_id"],{})
        rows.append({
            "week_start":   t["week_start"],
            "month_label":  t.get("month_label",""),
            "am1":          t.get("am1",""),
            "am2":          t.get("am2",""),
            "project_name": p.get("name",""),
            "task_col":     t["task_col"],
            "done":         "✅" if t.get("status","done" if t.get("checked") else "pending")=="done" else ("➖" if t.get("status")=="na" else "❌"),
            "updated":      str(t.get("updated_at",""))[:16],
        })
    df=pd.DataFrame(rows).sort_values(["week_start","am1","am2"],ascending=[False,True,True])
    fdf,_=filter_bar(df, show_week=True, show_month=True,
                     show_am1=True, show_am2=True, show_project=True, key_prefix="all_tasks")
    tc_f=st.selectbox("🔧 Task Type",["All"]+TASK_COLS, key="tc_f")
    st_f=st.selectbox("✅ Status",["All","Done","Pending"], key="st_f")
    if tc_f!="All": fdf=fdf[fdf["task_col"]==tc_f]
    if st_f=="Done":    fdf=fdf[fdf["done"]=="✅"]
    elif st_f=="Pending": fdf=fdf[fdf["done"]=="❌"]
    disp=fdf.copy()
    disp["week_start"]=disp["week_start"].apply(week_label)
    disp.columns=["Week","Month","AM1","AM2","Project","Task","Done","Updated"]
    st.dataframe(disp, use_container_width=True, hide_index=True, height=480)
    st.caption(f"{len(disp)} records | Total: {len(df)}")
    st.download_button("📥 Download CSV",fdf.to_csv(index=False).encode(),
                       "tasks_export.csv","text/csv")


def mgr_manage():
    st.markdown('<div class="ph"><h1>⚙️ Manage</h1></div>', unsafe_allow_html=True)
    tab1,tab2 = st.tabs(["👤 Users","🏗️ Projects"])
    with tab1:
        users=get_users()
        rc_map={"manager":"#3b82f6","executive":"#8b5cf6"}
        st.markdown("### Current Users")
        for u,d in users.items():
            rc=rc_map.get(d["role"],"#64748b")
            st.markdown(f"""
            <div class="card" style="display:flex;justify-content:space-between;align-items:center">
                <div><b>{d['name']}</b><span style="color:#64748b;font-size:.8rem"> @{u}</span></div>
                <span style="background:{rc};color:white;padding:2px 10px;
                      border-radius:20px;font-size:.72rem">{d['role'].upper()}</span>
            </div>""", unsafe_allow_html=True)
        st.markdown("---")
        st.markdown("### Add New Login")
        with st.form("uf"):
            c1,c2=st.columns(2)
            nm=c1.text_input("Full Name"); un=c2.text_input("Username")
            c3,c4=st.columns(2)
            pw=c3.text_input("Password",type="password")
            role=c4.selectbox("Role",["executive","manager"])
            if st.form_submit_button("Add User",type="primary"):
                if not nm or not un or not pw: st.error("All fields required")
                elif un in users:              st.error("Username exists")
                else:
                    db_insert("users",{"username":un,"name":nm,"role":role,"password":hp(pw)})
                    st.success(f"✅ {nm} added!"); st.rerun()
    with tab2:
        projects=get_projects()
        st.markdown("### Current Projects")
        if projects:
            st.dataframe(pd.DataFrame([{"Project":p["name"],"AM1":p["am1"],"AM2":p["am2"]}
                                        for p in projects]),
                         use_container_width=True, hide_index=True)
        st.markdown("---")
        st.markdown("### Add Project")
        e_am1=sorted({p["am1"] for p in projects}) if projects else []
        e_am2=sorted({p["am2"] for p in projects}) if projects else []
        with st.form("pf"):
            pname=st.text_input("Project Name")
            c1,c2=st.columns(2)
            am1_opts=e_am1+["+ New AM1..."]; am1_sel=c1.selectbox("AM1",am1_opts) if e_am1 else None
            am1_new=c1.text_input("New AM1 Name") if (not e_am1 or am1_sel=="+ New AM1...") else ""
            am2_opts=e_am2+["+ New AM2..."]; am2_sel=c2.selectbox("AM2",am2_opts) if e_am2 else None
            am2_new=c2.text_input("New AM2 Name") if (not e_am2 or am2_sel=="+ New AM2...") else ""
            if st.form_submit_button("Add Project",type="primary"):
                am1f=am1_new.strip() if (not e_am1 or am1_sel=="+ New AM1...") else am1_sel
                am2f=am2_new.strip() if (not e_am2 or am2_sel=="+ New AM2...") else am2_sel
                if not pname or not am1f or not am2f: st.error("All fields required")
                else:
                    add_project(pname,am1f,am2f)
                    st.success(f"✅ '{pname}' added!"); st.rerun()
        if projects:
            st.markdown("### Remove Project")
            opts={p["name"]:p["id"] for p in projects}
            del_n=st.selectbox("Select to remove",list(opts.keys()),key="del_p")
            if st.button("🗑️ Delete",type="secondary"):
                db_delete("projects","id", opts[del_n])
                st.success("Deleted."); st.rerun()

        if projects:
            st.markdown("---")
            st.markdown("### ✏️ Edit Project")
            proj_map = {p["name"]: p for p in projects}
            edit_name = st.selectbox("Select project to edit",
                                     list(proj_map.keys()), key="edit_proj_sel")
            sel_proj  = proj_map[edit_name]

            e_am1 = sorted({p["am1"] for p in projects})
            e_am2 = sorted({p["am2"] for p in projects})

            with st.form("edit_pf"):
                st.markdown(f"**Editing:** {sel_proj['name']}")
                c1, c2, c3 = st.columns(3)

                new_pname = c1.text_input("Project Name", value=sel_proj["name"])

                # AM1 — pick existing or type new
                am1_opts  = e_am1 + ["+ New AM1..."]
                am1_idx   = e_am1.index(sel_proj["am1"]) if sel_proj["am1"] in e_am1 else 0
                am1_sel   = c2.selectbox("AM1", am1_opts, index=am1_idx, key="edit_am1_sel")
                am1_new   = c2.text_input("New AM1 Name", key="edit_am1_new") \
                            if am1_sel == "+ New AM1..." else ""

                # AM2 — pick existing or type new
                am2_opts  = e_am2 + ["+ New AM2..."]
                am2_idx   = e_am2.index(sel_proj["am2"]) if sel_proj["am2"] in e_am2 else 0
                am2_sel   = c3.selectbox("AM2", am2_opts, index=am2_idx, key="edit_am2_sel")
                am2_new   = c3.text_input("New AM2 Name", key="edit_am2_new") \
                            if am2_sel == "+ New AM2..." else ""

                if st.form_submit_button("💾 Save Changes", type="primary"):
                    am1f = am1_new.strip() if am1_sel == "+ New AM1..." else am1_sel
                    am2f = am2_new.strip() if am2_sel == "+ New AM2..." else am2_sel
                    if not new_pname or not am1f or not am2f:
                        st.error("All fields required")
                    else:
                        db_update("projects", "id", sel_proj['id'],
                                  {"name": new_pname, "am1": am1f, "am2": am2f})
                        st.success(f"✅ '{new_pname}' updated!"); st.rerun()

# ─────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────
def main():
    css()
    try:
        seed()
    except Exception:
        pass

    if not st.session_state.get("logged_in"):
        login(); return

    role = st.session_state["role"]
    page = sidebar(role)

    if role=="executive":
        if   "Overview" in page: exec_overview()
        elif "Mark"     in page: exec_mark_tasks()
        elif "Weekly"   in page: exec_weekly_page()
        elif "Monthly"  in page: exec_monthly_page()
    elif role=="manager":
        if   "Dashboard" in page: mgr_dashboard()
        elif "Weekly"    in page: mgr_weekly_report()
        elif "Monthly"   in page: mgr_monthly_report()
        elif "History"   in page: mgr_report_history()
        elif "All Tasks" in page: mgr_all_tasks()
        elif "Manage"    in page: mgr_manage()

if __name__=="__main__":
    main()
